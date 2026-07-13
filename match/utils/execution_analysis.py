#!/usr/bin/env python3
"""
Integration module for extracting execution analysis and generating Excel reports.
Filters operations only (not constants/inputs), extracts detailed parameters from nodes.
For MATCH nodes: uses IR graph via ir_mod.
For TVM nodes: uses func_name to infer operation type and compute MACs.
"""

import json
import numpy as np
import re
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class NodeAnalyzer:
    """Analyzes model nodes and computes execution metrics."""
    
    def __init__(self, mod_info: Dict[str, Any], match_nodes_data: Optional[Dict[str, Any]] = None):
        """
        Initialize analyzer with model info.
        
        Args:
            mod_info: TVM mod_info dictionary from graph.py
            match_nodes_data: Dictionary with MATCH node info including ir_mod and schedule
        """
        self.mod_info = mod_info
        self.nodes = mod_info.get("nodes", [])
        self.shapes = mod_info.get("attrs", {}).get("shape", [None])[1] if "attrs" in mod_info else []
        self.dtypes = mod_info.get("attrs", {}).get("dltype", [None])[1] if "attrs" in mod_info else []
        self.heads = [head[0] for head in mod_info.get("heads", [])]
        self.match_nodes_data = match_nodes_data or {}
    
    def _find_conv_node_in_ir(self, ir_mod: Any) -> Optional[Any]:
        """
        Find and return the first Conv node in the IR module.
        
        Args:
            ir_mod: IR module from host_lib.ir_mod
            
        Returns:
            Conv node object or None if not found
        """
        try:
            if ir_mod is None:
                return None
            
            main_func = ir_mod["main"]
            if not hasattr(main_func, 'body'):
                return None
            
            # Traverse the function body to find Conv nodes
            body = main_func.body
            has_conv = "conv2d" in str(body).lower()
            if has_conv:
                conv = body
                while conv and conv.op.name != "nn.conv2d":
                    conv = conv.args[0] if hasattr(conv, 'args') and len(conv.args) > 0 else None
                    if not hasattr(conv, 'op'):
                        return None
                return conv
        except Exception as e:
            pass
        
        return None
    
    def get_analysis_data(self) -> List[Dict[str, Any]]:
        """
        Get complete analysis data for operation nodes only (skips constants and inputs).
        Filters out null ops and nop operations.
        
        Returns:
            List of dicts with node analysis data in execution order
        """
        analysis_data = []
        
        for exec_order, node in enumerate(self.nodes, 1):
            node_id = exec_order - 1
            
            if node_id >= len(self.shapes):
                continue
            
            # FILTER: Skip null ops (inputs and constants)
            op_type = node.get("op", "")
            if op_type == "null":
                continue
            
            # FILTER: Skip NOP (no-op/identity) operations
            node_name = node.get("name", f"node_{node_id}")
            if "_nop" in node_name:
                continue
            
            # Get shapes
            output_shape = tuple(self.shapes[node_id]) if node_id < len(self.shapes) else ()
            output_dtype = str(self.dtypes[node_id]) if node_id < len(self.dtypes) else ""
            
            input_shapes = []
            for inp in node.get("inputs", []):
                inp_node_id = inp[0]
                if inp_node_id < len(self.shapes):
                    input_shapes.append(tuple(self.shapes[inp_node_id]))
            
            # Get operation type from func_name or MATCH IR graph
            func_name = node.get("attrs", {}).get("func_name", "")
            node_name_lower = node_name.lower()
            func_name_lower = func_name.lower()
            

            output_channels = 1
            output_height = 1
            output_width = 1
            input_channels = 1
            kernel_height = 1
            kernel_width = 1
            data_layout = node.get("attrs", {}).get("data_layout", "") or node.get("attrs", {}).get("layout", "") or "NHWC"
            kernel_layout = node.get("attrs", {}).get("kernel_layout", "") or "HWIO"
            try:
                if data_layout=="NHWC":
                    output_height = output_shape[1]
                    output_width = output_shape[2]
                    output_channels = output_shape[3]
                    if len(input_shapes) > 0:
                        input_channels = input_shapes[0][3]
                    else:
                        input_channels = input_shapes[3]
            except Exception as e:
                pass

            if "match" in node_name_lower and node_name in self.match_nodes_data:
                match_data = self.match_nodes_data[node_name]
                ir_mod = match_data.get("ir_mod")
                conv_node = self._find_conv_node_in_ir(ir_mod) if ir_mod else None
                if conv_node:
                    kernel_size = dict(conv_node.attrs).get("kernel_size", (1, 1))
                    kernel_height, kernel_width = int(kernel_size[0]), int(kernel_size[1])
            else:
                if "conv" in node_name_lower or "conv" in func_name_lower:
                    if kernel_layout == "HWIO":
                        kernel_height = input_shapes[1][0]
                        kernel_width = input_shapes[1][1]

            macs = output_channels * output_height * output_width * input_channels * kernel_height * kernel_width
            
            analysis_data.append([
                node_id,
                "_".join(node_name.split("_")[2:]),
                func_name,
                str(input_shapes),
                str(output_shape),
                str(output_dtype),
                output_channels,
                output_height,
                output_width,
                input_channels,
                kernel_height,
                kernel_width,
                macs,
            ])
        
        return analysis_data


class ExecutionAnalysisExporter:
    """Exports execution analysis to Excel format."""
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    def _get_header_style(self):
        """Get header cell style."""
        return {
            "fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            "font": Font(bold=True, color="FFFFFF"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            ),
        }
    
    def _get_data_style(self):
        """Get data cell style."""
        return {
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            ),
        }
    
    def export_to_excel(self, analysis_data: List[List[Any]], output_file: str = "model_analysis.xlsx"):
        """
        Export analysis data to Excel file with separated shape components.
        Columns: Execution Order, Node ID, Node Name, Operation, Function Name,
                 Output Channels, Output Height, Output Width, Input Channels,
                 Kernel Height, Kernel Width, Strides, Padding, Groups/Units, Data Layout,
                 Input Shapes, Output Shape, Output Dtype, MACs, Output Node
        
        Args:
            analysis_data: List of dicts from NodeAnalyzer.get_analysis_data()
            output_file: Output file path
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Node Analysis"
        
        header_style = self._get_header_style()
        data_style = self._get_data_style()
        
        # Define headers with semantic shape components
        headers = [
            "Node ID",
            "Node Name",
            "Fname",
            "Inp Shapes",
            "Out Shape",
            "Out Dtype",
            "OC",
            "OH",
            "OW",
            "IC",
            "KH",
            "KW",
            "MACs",
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_style["fill"]
            cell.font = header_style["font"]
            cell.alignment = header_style["alignment"]
            cell.border = header_style["border"]
        
        # Write data rows
        for row_idx, row_data in enumerate(analysis_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = data_style["border"]
        
        # Set column widths
        col_widths = [10, 20, 20, 25, 25, 12, 10, 10, 10, 10, 10, 10, 15]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save workbook
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)


def generate_execution_report(mod_info: Dict[str, Any], output_file: str = "execution_analysis.xlsx",
                              match_nodes_data: Optional[Dict[str, Any]] = None) -> bool:
    """
    Generate execution analysis report in one command.
    Convenience function for quick usage.
    
    Filters operations only (skips inputs and constants).
    Extracts detailed parameters and breaks shapes into components.
    
    Args:
        mod_info: TVM mod_info dictionary from graph.py
        output_file: Output Excel file path
        match_nodes_data: Optional dictionary with MATCH node IR graph info
        
    Returns:
        True if successful, False otherwise
    """
    try:
        analyzer = NodeAnalyzer(mod_info, match_nodes_data)
        analysis_data = analyzer.get_analysis_data()
        
        if not analysis_data:
            print("[EXECUTION ANALYSIS] Warning: No operation nodes found")
            return False
        
        exporter = ExecutionAnalysisExporter()
        exporter.export_to_excel(analysis_data, output_file)
        
        return True
    except ImportError:
        raise
    except Exception as e:
        print(f"[EXECUTION ANALYSIS] Error generating report: {e}")
        import traceback
        traceback.print_exc()
        return False
